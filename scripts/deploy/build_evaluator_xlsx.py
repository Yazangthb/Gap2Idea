"""
Build deploy/gap2idea_evaluator.xlsx from the three pipeline runs.

Two sheets:
  - "Instructions"      one-page brief for evaluators
  - "Ideas & Scoring"   15 rows, one per idea; evaluators fill the last 4 cols
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO = Path(__file__).resolve().parents[2]
OUT = REPO / "deploy" / "gap2idea_evaluator_v2.xlsx"
DOMAIN_LABELS = {"ai": "AI", "ml": "Classical ML", "math": "Mathematics"}

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
SCORE_FILL = PatternFill("solid", start_color="FFF2CC")
DOMAIN_FILL = {
    "AI": PatternFill("solid", start_color="DDEBF7"),
    "Classical ML": PatternFill("solid", start_color="E2EFDA"),
    "Mathematics": PatternFill("solid", start_color="FCE4D6"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def load_ideas():
    """
    For each domain: read ideas.tsv + idea_eval.tsv, compute the judge
    composite (mean of novelty / specificity / feasibility / grounding),
    take the top 5. If idea_eval.tsv is missing or fewer than 5 rows have
    eval data, fall back to first-5 of ideas.tsv.
    """
    AXES = ["novelty", "specificity", "feasibility", "evidence_grounding"]
    rows, n = [], 0
    for key in ("ai", "ml", "math"):
        ideas = pd.read_csv(REPO / "runs" / key / "artifacts" / "ideas.tsv", sep="\t")
        eval_path = REPO / "runs" / key / "artifacts" / "idea_eval.tsv"
        if eval_path.exists():
            evals = pd.read_csv(eval_path, sep="\t")
            m = ideas.merge(
                evals[["title"] + AXES],
                on="title", how="left",
            )
            m["composite"] = m[AXES].mean(axis=1)
            m = m.sort_values("composite", ascending=False).reset_index(drop=True)
            picked = m.head(5)
        else:
            picked = ideas.head(5)
        for _, r in picked.iterrows():
            n += 1
            rows.append({
                "n": n,
                "domain": DOMAIN_LABELS[key],
                "title": str(r["title"]).strip(),
                "rq": str(r["research_question"]).strip(),
                "method": str(r["method_sketch"]).strip(),
                "evalp": str(r["evaluation_plan"]).strip(),
                "risks": str(r["assumptions_and_risks"]).strip(),
            })
    return rows


def main():
    ideas = load_ideas()

    wb = Workbook()

    # ----- Sheet 1: Instructions -----
    ws1 = wb.active
    ws1.title = "Instructions"
    ws1["A1"] = "Gap2Idea — Research Idea Evaluation"
    ws1["A1"].font = Font(name=FONT, bold=True, size=16)
    ws1.row_dimensions[1].height = 24

    intro = [
        "",
        "Thank you for evaluating these research ideas.",
        "",
        "WHAT TO DO",
        "  1. Open the 'Ideas & Scoring' sheet.",
        "  2. For each of the 15 ideas, read the title, research question, "
        "method sketch, evaluation plan, and risks.",
        "  3. Score each idea on three 1-5 Likert axes:",
        "       Novelty           1 = not novel        5 = highly novel",
        "       Feasibility       1 = very hard         5 = clearly feasible",
        "       Potential impact  1 = marginal         5 = field-shifting",
        "  4. Leave a free-text comment if you want (optional).",
        "",
        "DOMAINS",
        "  Ideas 1-5    AI                (cs.CL / cs.LG / cs.AI papers)",
        "  Ideas 6-10   Classical ML      (stat.ML / cs.LG papers)",
        "  Ideas 11-15  Mathematics       (math.OC / PR / AP / FA papers)",
        "",
        "TIME",
        "  Expected: 12-15 minutes.",
        "",
        "HOW THESE WERE GENERATED",
        "  Each idea was synthesised by the Gap2Idea pipeline from a 25-paper "
        "arXiv subset per domain. The LLM extracted limitations / future-work "
        "sentences, clustered them into themes, and produced bridge-style ideas "
        "spanning two themes. See https://github.com/Yazangthb/Gap2Idea for the "
        "method.",
        "",
        "SCORES ARE PRE-FILLED WHERE? Nowhere — every score cell is empty for you.",
    ]
    for i, line in enumerate(intro, start=2):
        cell = ws1.cell(row=i, column=1, value=line)
        cell.font = Font(name=FONT, size=11, bold=line.isupper() and len(line) > 0)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws1.column_dimensions["A"].width = 110

    # ----- Sheet 2: Ideas & Scoring -----
    ws = wb.create_sheet("Ideas & Scoring")

    headers = [
        "#", "Domain", "Title",
        "Research question", "Method sketch", "Evaluation plan",
        "Assumptions & risks",
        "Novelty (1-5)", "Feasibility (1-5)", "Potential impact (1-5)",
        "Comments",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = BORDER
    ws.row_dimensions[1].height = 32

    # Data rows
    for r, idea in enumerate(ideas, start=2):
        ws.cell(row=r, column=1, value=idea["n"]).alignment = Alignment(
            horizontal="center", vertical="top")
        ws.cell(row=r, column=2, value=idea["domain"]).alignment = Alignment(
            horizontal="center", vertical="top")
        ws.cell(row=r, column=3, value=idea["title"])
        ws.cell(row=r, column=4, value=idea["rq"])
        ws.cell(row=r, column=5, value=idea["method"])
        ws.cell(row=r, column=6, value=idea["evalp"])
        ws.cell(row=r, column=7, value=idea["risks"])
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=r, column=col)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(
                wrap_text=True, vertical="top",
                horizontal="center" if col in (1, 2, 8, 9, 10) else "left",
            )
            c.border = BORDER
            if col == 2:  # Domain tint on the domain cell
                c.fill = DOMAIN_FILL[idea["domain"]]
            if col in (8, 9, 10):  # Score columns
                c.fill = SCORE_FILL
        ws.row_dimensions[r].height = 200  # plenty of room for the long text

    # Summary footer with formulas
    last_idea_row = 1 + len(ideas)
    summary_row = last_idea_row + 2
    ws.cell(row=summary_row, column=7, value="MEAN").font = Font(
        name=FONT, bold=True, size=11)
    ws.cell(row=summary_row, column=7).alignment = Alignment(horizontal="right")
    for col in (8, 9, 10):
        col_letter = get_column_letter(col)
        formula = f"=IFERROR(AVERAGE({col_letter}2:{col_letter}{last_idea_row}),\"\")"
        c = ws.cell(row=summary_row, column=col, value=formula)
        c.font = Font(name=FONT, bold=True, size=11)
        c.alignment = Alignment(horizontal="center")
        c.fill = SCORE_FILL
        c.number_format = "0.00"

    # Column widths
    widths = {
        1: 5, 2: 14, 3: 32,
        4: 45, 5: 50, 6: 45, 7: 40,
        8: 14, 9: 14, 10: 18, 11: 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze header + first 3 cols (#, Domain, Title) so the long text stays
    # visible while scrolling right.
    ws.freeze_panes = "D2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}  ({len(ideas)} ideas)")


if __name__ == "__main__":
    main()
